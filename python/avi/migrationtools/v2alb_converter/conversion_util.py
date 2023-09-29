# Copyright 2021 VMware, Inc.
# SPDX-License-Identifier: Apache License 2.0
import copy
import logging
import os
import random
import re
from functools import reduce

import avi.migrationtools.f5_converter.converter_constants as conv_const
import pandas
from avi.migrationtools.avi_migration_utils import MigrationUtil
from openpyxl import load_workbook
from pkg_resources import parse_version
from xlsxwriter import Workbook

LOG = logging.getLogger(__name__)
csv_writer_dict_list = []

# Added variable for checking progress and get overall object.
ppcount = 0
ptotal_count = 0
global fully_migrated
fully_migrated = 0
used_pool_groups = {}
used_pool = {}


class NsxvConvUtil(MigrationUtil):
    def generate_id_for_converted_objects(self, edge_name, obj_id, obj_suffix, prefix):
        obj_name = f"{edge_name}-{obj_id}-{obj_suffix}"
        if prefix:
            obj_name = f"{prefix}-{obj_name}"
        return obj_name

    def get_total_objects_count_in_v(self, v_edge_config, obj_key):
        total_obj = 0
        for edge_name, edge_config in v_edge_config.items():
            if edge_config.get(obj_key):
                total_obj += len(edge_config.get(obj_key))

        LOG.debug(f"get_total_objects_count_in_v: {total_obj}")
        return total_obj

    def add_conv_status(
        self, nsxv_type, nsxv_sub_type, nsxv_id, conv_status, avi_object=None
    ):

        global csv_writer_dict_list
        # Added space if NsxV_sub_type None for pivot table
        row = {
            "NsxV type": nsxv_type,
            "NsxV SubType": nsxv_sub_type if nsxv_sub_type else " ",
            "NsxV ID": nsxv_id,
            "Status": conv_status.get("status", ""),
            "Skipped settings": str(conv_status.get("skipped", "")),
            "Indirect mapping": str(conv_status.get("indirect", "")),
            "Not Applicable": str(conv_status.get("na_list", "")),
            "Avi Object": str(avi_object),
        }
        csv_writer_dict_list.append(row)

    def add_complete_conv_status(
        self, output_dir, avi_config, report_name, vs_level_status
    ):

        global csv_writer_dict_list
        global ptotal_count
        for status in conv_const.STATUS_LIST:
            status_list = [
                row for row in csv_writer_dict_list if row["Status"] == status
            ]
            if not status == "ERROR":
                print(f"Total {status} OBJECT: {len(status_list)}")
                LOG.debug(f"Total {status} OBJECT: {len(status_list)}")
            else:
                print(f"Total {status} COUNT: {len(status_list)}")
                LOG.debug(f"Total {status} COUNT: {len(status_list)}")

        msg = "Writing Excel Sheet For Converted Configuration..."
        print(msg)
        LOG.debug(msg)
        ptotal_count = ptotal_count + len(csv_writer_dict_list)
        if vs_level_status:
            self.vs_per_skipped_setting_for_references(avi_config)
            self.correct_vs_ref(avi_config)
        else:
            # Update the complexity level of VS as Basic or Advanced
            self.vs_complexity_level()
        self.write_status_report_and_pivot_table_in_xlsx(
            output_dir, report_name, vs_level_status
        )

    def write_status_report_and_pivot_table_in_xlsx(
        self, output_dir, report_name, vs_level_status
    ):
        """
        This function defines that add status sheet and pivot table sheet in
        xlsx format
        :param output_dir: Path of output directory
        :param report_name: filename to write report
        :param vs_level_status: Flag to include VS wise detailed status or not
        :return: None
        """
        global ppcount
        global ptotal_count
        # List of fieldnames for headers
        if vs_level_status:
            fieldnames = [
                "NsxV type",
                "NsxV SubType",
                "NsxV ID",
                "Status",
                "Skipped settings",
                "Indirect mapping",
                "Not Applicable",
                'Complexity Level',
                "VS Reference",
                "Overall skipped settings",
                "Avi Object",
            ]
        else:
            fieldnames = [
                "NsxV type",
                "NsxV SubType",
                "NsxV ID",
                "Status",
                "Skipped settings",
                "Indirect mapping",
                "Not Applicable",
                "Avi Object",
            ]
        LOG.debug(f"Fieldnames for headers {fieldnames}")

        # xlsx workbook
        report_path = (
            output_dir + os.path.sep + "%s-ConversionStatus.xlsx" % report_name
        )
        status_wb = Workbook(report_path)
        # xlsx worksheet
        status_ws = status_wb.add_worksheet("Status Sheet")
        LOG.debug("Added worksheet in workbook")
        # Lock the first row of xls report.
        status_ws.freeze_panes(1, 0)
        first_row = 0
        for header in fieldnames:
            col = fieldnames.index(header)
            status_ws.write(first_row, col, header)
        row = 1
       
        for row_data in csv_writer_dict_list:
            ppcount += 1
            for _key, _value in row_data.items():
                col = fieldnames.index(_key)
                status_ws.write(row, col, _value)
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            LOG.debug(msg)
            self.print_progress_bar(
                ppcount, ptotal_count, msg, prefix="Progress", suffix=""
            )
            row += 1
        status_wb.close()
        # create dataframe for row list
        df = pandas.DataFrame(csv_writer_dict_list, columns=fieldnames)
        # create pivot table using pandas
        pivot_table = pandas.pivot_table(
            df,
            index=["Status", "NsxV type", "NsxV SubType"],
            values=[],
            aggfunc=[len],
            fill_value=0,
        )

        # create dataframe for pivot table using pandas
        pivot_df = pandas.DataFrame(pivot_table)
        LOG.debug("Created dataframe...")
        main_book = load_workbook(report_path)
        main_writer = pandas.ExcelWriter(
            report_path, engine="openpyxl", mode="a")
        main_writer._book = main_book

        # Add pivot table in Pivot sheet
        pivot_df.to_excel(main_writer, "Pivot Sheet")
        LOG.debug("Added pivot table in pivot sheet...")
        main_writer.close()

    def add_status_row(self, nsxv_type, nsxv_sub_type, nsxv_id, status, avi_obj=None):
        """
        Adds as status row in conversion status csv
        :param nsxv_type: Object type
        :param nsxv_sub_type: Object sub type
        :param nsxv_id: Name of object
        :param status: conversion status
        :param avi_obj: Converted avi object
        """
        global csv_writer_dict_list
        # Added space if nsxv_sub_type None for pivot table
        row = {
            "NsxV type": nsxv_type,
            "NsxV SubType": nsxv_sub_type if nsxv_sub_type else " ",
            "NsxV ID": nsxv_id,
            "Status": status,
        }
        if avi_obj:
            row.update({"Avi Object": str(avi_obj)})
        csv_writer_dict_list.append(row)

    def vs_complexity_level(self):
        """
        This method calculate the complexity of vs.
        :return:
        """
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [
            row
            for row in csv_writer_dict_list
            if row["Status"]
            in [conv_const.STATUS_PARTIAL, conv_const.STATUS_SUCCESSFUL]
            and row["NsxV type"] == "virtual"
        ]
        for vs_csv_object in vs_csv_objects:
            virtual_service = self.format_string_to_json(
                vs_csv_object["Avi Object"])
            # Update the complexity level of VS as Basic or Advanced
            LOG.debug("Updating complexity level of VS")
            self.update_vs_complexity_level(vs_csv_object, virtual_service)

    def vs_per_skipped_setting_for_references(self, avi_config):
        """
        This functions defines that Add the skipped setting per VS CSV row
        :param avi_config: this method use avi_config for checking vs skipped
        :return: None
        """

        # Get the count of vs fully migrated
        global fully_migrated
        global ptotal_count
        global ppcount
        fully_migrated = 0
        # Get the VS object list which is having status successful and partial.
        LOG.debug("Get the VS object list which is having status successful and partial.")
        vs_csv_objects = [
            row
            for row in csv_writer_dict_list
            if row["Status"]
            in [conv_const.STATUS_PARTIAL, conv_const.STATUS_SUCCESSFUL]
            and row["NsxV type"] == "virtualservice"
        ]
        # Get the list of csv rows which has profile as NsxV type
        LOG.debug("Get the list of csv rows which has profile as NsxV type")
        profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ["applicationprofile"]
        )
        ssl_profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ["sslprofile"]
        )
        ssl_key_certificate_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ["ssl_key_and_certificate"]
        )
        ptotal_count = ptotal_count + len(vs_csv_objects)
        for vs_csv_object in vs_csv_objects:
            ppcount += 1
            skipped_setting = {}
            virtual_service = self.format_string_to_json(
                vs_csv_object["Avi Object"])
            # Update the complexity level of VS as Basic or Advanced
            LOG.debug("Update the complexity level of VS as Basic or Advanced")
            self.update_vs_complexity_level(vs_csv_object, virtual_service)
            vs_ref = virtual_service["name"]
            repls = ("[", ""), ("]", "")
            # Get list of skipped setting attributes
            LOG.debug("Get list of skipped setting attributes")
            skipped_setting_csv = reduce(
                lambda a, kv: a.replace(
                    *kv), repls, vs_csv_object["Skipped settings"]
            )
            if skipped_setting_csv:
                skipped_setting["virtual_service"] = [skipped_setting_csv]
            # Get the skipped list for ssl key and cert
            LOG.debug("Get the skipped list for ssl key and cert")
            if "ssl_key_and_certificate_refs" in virtual_service:
                for ssl_key_and_certificate_ref in virtual_service[
                    "ssl_key_and_certificate_refs"
                ]:
                    ssl_key_cert = self.get_name(ssl_key_and_certificate_ref)
                    ssl_kc_skip = self.get_csv_skipped_list(
                        ssl_key_certificate_csv_list,
                        ssl_key_cert,
                        vs_ref,
                        field_key="ssl_cert_key",
                    )
                    if ssl_kc_skip:
                        skipped_setting["ssl cert key"] = {}
                        skipped_setting["ssl cert key"]["name"] = ssl_key_cert
                        skipped_setting["ssl cert key"]["skipped_list"] = ssl_kc_skip

            # Get the skipped list for ssl profile name.
            # Changed ssl profile name to ssl profile ref.
            LOG.debug("Get the skipped list for ssl profile name")
            if "ssl_profile_ref" in virtual_service:
                name, skipped = self.get_ssl_profile_skipped(
                    ssl_profile_csv_list, virtual_service["ssl_profile_ref"], vs_ref
                )
                if skipped:
                    skipped_setting["ssl profile"] = {}
                    skipped_setting["ssl profile"]["name"] = name
                    skipped_setting["ssl profile"]["skipped_list"] = skipped
            # Get the skipped list for pool group.
            LOG.debug("Get the skipped list for pool group.")
            if "pool_group_ref" in virtual_service:
                pool_group_name = self.get_name(
                    virtual_service["pool_group_ref"])
                csv_pool_rows = self.get_csv_object_list(
                    csv_writer_dict_list, ["pool"])
                pool_group_skipped_settings = self.get_pool_skipped_list(
                    avi_config,
                    pool_group_name,
                    csv_pool_rows,
                    csv_writer_dict_list,
                    vs_ref,
                    profile_csv_list,
                )
                if pool_group_skipped_settings:
                    skipped_setting["Pool Group"] = pool_group_skipped_settings
            # Get the skipped list for pool.
            LOG.debug("Get the skipped list for pool.")
            if "pool_ref" in virtual_service:
                pool_skipped_settings = {"pools": []}
                pool_name = self.get_name(virtual_service["pool_ref"])
                csv_pool_rows = self.get_csv_object_list(
                    csv_writer_dict_list, ["pool"])
                self.get_skipped_pool(
                    avi_config,
                    pool_name,
                    csv_pool_rows,
                    csv_writer_dict_list,
                    vs_ref,
                    profile_csv_list,
                    pool_skipped_settings,
                )
                if pool_skipped_settings["pools"]:
                    skipped_setting["Pool"] = pool_skipped_settings
            # Get the skipepd list for http policy.
            LOG.debug("Get the skipepd list for http policy.")
            if "http_policies" in virtual_service:
                policy_csv_list = self.get_csv_object_list(
                    csv_writer_dict_list, ["policy", "profile"]
                )
                for http_ref in virtual_service["http_policies"]:
                    policy_set_name, skipped_list = self.get_policy_set_skipped(
                        policy_csv_list, http_ref["http_policy_set_ref"], vs_ref
                    )
                    if skipped_list:
                        skipped_setting["Httppolicy"] = {}
                        skipped_setting["Httppolicy"]["name"] = policy_set_name
                        skipped_setting["Httppolicy"]["skipped_list"] = skipped_list
                    # Get the http policy name
                    LOG.debug("Get the http policy name")
                    pool_csv_rows = self.get_csv_object_list(
                        csv_writer_dict_list, ["pool"]
                    )
                    for each_http_policy in avi_config["HTTPPolicySet"]:
                        if (
                            each_http_policy["name"] == policy_set_name
                            and "http_request_policy" in each_http_policy
                        ):
                            for http_req in each_http_policy["http_request_policy"][
                                "rules"
                            ]:
                                if http_req.get("switching_action", {}):
                                    self.get_skip_pools_policy(
                                        policy_set_name,
                                        http_req,
                                        avi_config,
                                        pool_csv_rows,
                                        vs_ref,
                                        profile_csv_list,
                                        skipped_setting,
                                    )

            # Get the skipped list for application_profile_ref.
            LOG.debug("Get the skipped list for application_profile_ref")
            if "application_profile_ref" in virtual_service:
                name, skipped = self.get_application_profile_skipped(
                    profile_csv_list, virtual_service["application_profile_ref"], vs_ref
                )
                if skipped:
                    skipped_setting["Application profile"] = {}
                    skipped_setting["Application profile"]["name"] = name
                    skipped_setting["Application profile"]["skipped_list"] = skipped
            # Get the skipped list for network profile ref.
            LOG.debug("Get the skipped list for network profile ref.")
            if (
                "network_profile_ref" in virtual_service
                and "admin:System" not in virtual_service["network_profile_ref"]
            ):
                name, skipped = self.get_network_profile_skipped(
                    profile_csv_list, virtual_service["network_profile_ref"], vs_ref
                )
                if skipped:
                    skipped_setting["Network profile"] = {}
                    skipped_setting["Network profile"]["name"] = name
                    skipped_setting["Network profile"]["skipped_list"] = skipped
            # Update overall skipped setting of VS csv row
            LOG.debug("Update overall skipped setting of VS csv row")
            if skipped_setting:
                vs_csv_object.update(
                    {"Overall skipped settings": str(skipped_setting)})
            else:
                vs_csv_object.update(
                    {"Overall skipped settings": "FULLY MIGRATION"})
                fully_migrated += 1
            # Added call for progress function.
            LOG.debug("Added call for progress function.")
            msg = "excel sheet conversion started..."
            self.print_progress_bar(
                ppcount, ptotal_count, msg, prefix="Progress", suffix=""
            )
        csv_objects = [
            row
            for row in csv_writer_dict_list
            if row["Status"]
            in [conv_const.STATUS_PARTIAL, conv_const.STATUS_SUCCESSFUL]
            and row["NsxV type"] != "virtualservice"
        ]

        # Update the vs reference not in used if objects are not attached to
        # VS directly or indirectly
        LOG.debug("Update the vs reference not in used if objects are not attached to VS directly or indirectly")
        for row in csv_objects:
            if "VS Reference" not in row or row["VS Reference"] == "":
                row["VS Reference"] = conv_const.STATUS_NOT_IN_USE

    def correct_vs_ref(self, avi_config):
        """
        This method corrects the reference of VS to different objects
        :param avi_config: avi configuration dict
        :return:
        """
        global csv_writer_dict_list
        avi_graph = self.make_graph(avi_config)
        csv_dict_sub = [
            row
            for row in csv_writer_dict_list
            if row["NsxV type"] != "virtualservice"
            and row["Status"]
            in (conv_const.STATUS_PARTIAL, conv_const.STATUS_SUCCESSFUL)
        ]
        LOG.debug("Adding correct vs ref")
        for dict_row in csv_dict_sub:
            obj = dict_row["Avi Object"]
            
            vs = []
            if obj.startswith("{"):
                obj = eval(obj)
                for key in obj:
                   
                    for objs in obj[key]:
                       
                        self.add_vs_ref(objs, avi_graph, vs)
            elif obj.startswith("["):
                obj = eval(obj)
               
                for objs in obj:
                    for key in objs:
                        objval = objs[key]
                        self.add_vs_ref(objval, avi_graph, vs)
            if vs:
                dict_row["VS Reference"] = str(list(set(vs)))
            else:
                dict_row["VS Reference"] = conv_const.STATUS_NOT_IN_USE

    def get_conv_status(
        self,
        skipped,
        indirect_list,
        ignore_dict,
        input_object,
        user_ignore=None,
        na_list=None,
    ):
        """
        Update skipped list for conversion status
        :param skipped: All skipped attributes after conversion
        :param indirect_list: List of attrs to be mapped as indirect mapping
        :param ignore_dict: Dict of default values for column skipped for defaults
        :param input_object: Currant input object
        :param user_ignore: List of attributes user wants not to be shown in skipped
        :param na_list: List of attributes marked as not applicable
        :return: Conversion status dict
        """
        LOG.debug("Get conversion status")
        conv_status = dict()
        user_ignore = [] if not user_ignore else user_ignore
        na_list = [] if not na_list else na_list

        conv_status["user_ignore"] = [
            val for val in skipped if val in user_ignore]
        skipped = [attr for attr in skipped if attr not in user_ignore]

        conv_status["indirect"] = [
            val for val in skipped if val in indirect_list]
        skipped = [attr for attr in skipped if attr not in indirect_list]

        conv_status["na_list"] = [val for val in skipped if val in na_list]
        skipped = [attr for attr in skipped if attr not in na_list]

        default_skip = []
        for key in ignore_dict.keys():
            val = input_object.get(key)
            default_val = ignore_dict.get(key)
            if key in skipped and val == default_val:
                default_skip.append(key)
        if default_skip:
            skipped = [attr for attr in skipped if attr not in default_skip]

        conv_status["skipped"] = skipped
        conv_status["default_skip"] = default_skip
        if skipped:
            status = conv_const.STATUS_PARTIAL
        else:
            status = conv_const.STATUS_SUCCESSFUL
        conv_status["status"] = status
        return conv_status


    def get_csv_object_list(self, csv_writer_dict_list, command_list):
        """
        This method is used for getting csv object
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param command_list: List of netscaler commands
        :return: List of CSV rows
        """

        csv_object = [row for row in csv_writer_dict_list if
                      row['Status'] in [conv_const.STATUS_PARTIAL,
                                        conv_const.STATUS_SUCCESSFUL] and
                      '->' not in row['Avi Object'] and
                      row['NsxV type'] in command_list]
        return csv_object
    
    def get_avi_obj_from_name(self, obj_name, avi_config_dict, obj_type):
        avi_obj = [
            obj for obj in avi_config_dict[obj_type] if obj.get("name") == obj_name
        ]
        if avi_obj:
            return avi_obj[0]
        return None
